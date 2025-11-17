"""Excel processing tools for parsing TBS Price Change Summary reports"""

import pandas as pd
from pathlib import Path
from typing import Any, Dict, List
from crewai.tools import BaseTool
from pydantic import BaseModel, Field
import openpyxl
import re
import os


class ExcelReaderInput(BaseModel):
    """Input schema for ExcelReaderTool"""
    file_path: str = Field(..., description="Path to the Excel file to read")
    skip_rows: int = Field(7, description="Number of rows to skip before header (default: 7 for TBS reports)")


class ExcelReaderTool(BaseTool):
    name: str = "Excel File Reader"
    description: str = (
        "Reads Excel files and extracts data. Specifically designed for TBS Price Change Summary reports. "
        "Handles complex headers, skips metadata rows, and returns structured data as a list of dictionaries."
    )
    args_schema: type[BaseModel] = ExcelReaderInput

    def _run(self, file_path: str, skip_rows: int = 7) -> str:
        """
        Read and parse Excel file

        Args:
            file_path: Path to Excel file
            skip_rows: Number of rows to skip (default 7 for TBS reports)

        Returns:
            JSON string with parsed data
        """
        try:
            # Read Excel file
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                return f"Error: File not found at {file_path}"

            # Read with pandas, skipping header rows and reading only the first 12 columns (A-L)
            # This avoids duplicate column names from None columns
            df = pd.read_excel(file_path, skiprows=skip_rows, usecols="A:L")

            # Clean column names (remove whitespace, newlines)
            df.columns = df.columns.str.strip().str.replace('\n', ' ')

            # Remove completely empty rows
            df = df.dropna(how='all')

            # Convert to list of dictionaries
            records = df.to_dict('records')

            # Build response
            result = {
                "success": True,
                "file_path": str(file_path),
                "total_records": len(records),
                "columns": list(df.columns),
                "records": records,  # Return all records (no longer limiting to preview)
                "full_record_count": len(records),
                "message": f"Successfully parsed {len(records)} records from Excel file"
            }

            import json
            return json.dumps(result, default=str, indent=2)

        except Exception as e:
            return f"Error reading Excel file: {str(e)}"

    def tool(self):
        """Return self for CrewAI compatibility"""
        return self


class DataCleanerInput(BaseModel):
    """Input schema for DataCleanerTool"""
    data: str = Field(..., description="JSON string of data to clean")


class DataCleanerTool(BaseTool):
    name: str = "Data Cleaner"
    description: str = (
        "Cleans and standardizes data. Handles missing values, normalizes text fields, "
        "converts data types, and removes special characters."
    )
    args_schema: type[BaseModel] = DataCleanerInput

    def _run(self, data: str) -> str:
        """
        Clean and standardize data

        Args:
            data: JSON string of data to clean

        Returns:
            JSON string with cleaned data
        """
        try:
            import json

            # Parse input data
            data_dict = json.loads(data)
            records = data_dict.get("records", [])

            cleaned_records = []
            for record in records:
                cleaned = {}
                for key, value in record.items():
                    # Handle None/NaN values
                    if pd.isna(value) or value is None:
                        cleaned[key] = None
                        continue

                    # Clean text fields
                    if isinstance(value, str):
                        # Strip whitespace
                        value = value.strip()
                        # Replace multiple spaces with single space
                        value = ' '.join(value.split())

                    cleaned[key] = value

                cleaned_records.append(cleaned)

            result = {
                "success": True,
                "cleaned_records": cleaned_records,
                "total_cleaned": len(cleaned_records),
                "message": f"Successfully cleaned {len(cleaned_records)} records"
            }

            return json.dumps(result, default=str, indent=2)

        except Exception as e:
            return f"Error cleaning data: {str(e)}"

    def tool(self):
        """Return self for CrewAI compatibility"""
        return self


class PriceCalculatorInput(BaseModel):
    """Input schema for PriceCalculatorTool"""
    old_price: float = Field(..., description="Old price value")
    new_price: float = Field(..., description="New price value")


class PriceCalculatorTool(BaseTool):
    name: str = "Price Calculator"
    description: str = (
        "Calculates price changes, percentage changes, and validates pricing data. "
        "Takes old and new prices and returns change amount, percentage, and direction."
    )
    args_schema: type[BaseModel] = PriceCalculatorInput

    def _run(self, old_price: float, new_price: float) -> str:
        """
        Calculate price change statistics

        Args:
            old_price: Original price
            new_price: New price

        Returns:
            JSON string with price change analysis
        """
        try:
            import json

            # Calculate changes
            change_amount = new_price - old_price

            if old_price != 0:
                change_percent = (change_amount / old_price) * 100
            else:
                change_percent = 0

            # Calculate price ratio for ±4% threshold categorization
            price_ratio = new_price / old_price if old_price != 0 else 1.0

            # Determine direction and type based on plan.md ±4% threshold rules:
            # - Permanent Change: 96% ≤ new price ≤ 104% of old price
            # - Begin LTO: new price < 96% of old price (>4% decrease)
            # - End LTO: new price > 104% of old price (>4% increase)
            if price_ratio < 0.96:
                direction = "decrease"
                change_type = "Begin LTO"
                is_significant = True
            elif price_ratio > 1.04:
                direction = "increase"
                change_type = "End LTO"
                is_significant = True
                # Note: "End LTO & Permanent Change" requires historical data check (future enhancement)
            elif 0.96 <= price_ratio <= 1.04:
                direction = "decrease" if change_amount < 0 else "increase"
                change_type = "Permanent Change"
                is_significant = True
            else:
                direction = "no change"
                change_type = "No Change"
                is_significant = False

            # Additional categorization based on Type of Sale field:
            # - If type_of_sale == "TBS - Licensee": change_type = "Licensee Change"
            # - If marked as new SKU in file: change_type = "New SKU"
            # (These would need to be passed as additional parameters to this function)

            result = {
                "success": True,
                "old_price": old_price,
                "new_price": new_price,
                "change_amount": round(change_amount, 2),
                "change_percent": round(change_percent, 2),
                "direction": direction,
                "suggested_category": change_type,
                "is_significant": is_significant,
                "formatted_change": f"${abs(change_amount):.2f}",
                "sign": "+" if change_amount >= 0 else "-"
            }

            return json.dumps(result, indent=2)

        except Exception as e:
            return f"Error calculating price change: {str(e)}"

    def tool(self):
        """Return self for CrewAI compatibility"""
        return self


class FormulaExcelGeneratorInput(BaseModel):
    """Input schema for FormulaExcelGeneratorTool"""
    input_file_path: str = Field(..., description="Path to the input Excel file")
    output_dir: str = Field(..., description="Directory where the formula Excel file will be saved")
    skip_rows: int = Field(7, description="Number of header rows to skip before data starts (default: 7 for TBS reports)")
    formula_column: str = Field("N", description="Column letter where formula should be added (default: N)")


class FormulaExcelGeneratorTool(BaseTool):
    name: str = "Formula Excel Generator"
    description: str = (
        "Creates a copy of the input Excel file with formulas added to generate formatted price change text. "
        "The formula combines product name, pack size, and price change information. "
        "Output file is saved with '_formula.xlsx' suffix using the same month/date from the input filename."
    )
    args_schema: type[BaseModel] = FormulaExcelGeneratorInput

    def _run(self, input_file_path: str, output_dir: str, skip_rows: int = 7, formula_column: str = "N") -> str:
        """
        Generate Excel file with formulas added

        Args:
            input_file_path: Path to input Excel file
            output_dir: Directory to save output file
            skip_rows: Number of header rows to skip (default 7)
            formula_column: Column to add formula (default N)

        Returns:
            JSON string with result information
        """
        try:
            import json
            from datetime import datetime

            # Validate input file exists
            input_path = Path(input_file_path)
            if not input_path.exists():
                return json.dumps({"success": False, "error": f"Input file not found: {input_file_path}"})

            # Extract month and date from input filename
            # Expected pattern: "TBS Price Change Summary Report - Month XXth'YY.xlsx"
            filename = input_path.stem  # Get filename without extension

            # Extract the date portion (e.g., "October 13th'25")
            match = re.search(r'-\s*(.+?)\.xlsx', input_path.name)
            if not match:
                # Fallback: try to extract just the date part
                match = re.search(r'([A-Za-z]+\s+\d{1,2}(?:st|nd|rd|th)\'?\d{2})', filename)

            if match:
                date_part = match.group(1).strip()
            else:
                # Fallback: use current date
                date_part = datetime.now().strftime("%B %d'%y")

            # Construct output filename
            output_filename = f"TBS Price Change Summary Report - {date_part}_formula.xlsx"
            output_path = Path(output_dir) / output_filename

            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)

            # Load the workbook
            wb = openpyxl.load_workbook(input_file_path)
            ws = wb.active

            # Calculate data start row (1-indexed for openpyxl)
            # skip_rows is the number of rows before headers, data typically starts 1 row after
            data_start_row = skip_rows + 2  # +1 for header row, +1 for first data row
            header_row = skip_rows + 1  # The row with column headers

            # Add header for column O (Price Ratio %)
            ws[f'O{header_row}'] = 'Price Ratio %'

            # Find the last row with data
            max_row = ws.max_row

            # Add formulas to each data row
            formula_count = 0
            for row_num in range(data_start_row, max_row + 1):
                # Check if row has data (check if column A has value)
                if ws[f'A{row_num}'].value is None:
                    continue

                # Formula: =PROPER($D{row})&" "&$H{row}&$L{row}&" "&$M{row}&"$"&ABS($K{row})&" to $"&$J{row}
                formula = f'=PROPER($D{row_num})&" "&$H{row_num}&$L{row_num}&" "&$M{row_num}&"$"&ABS($K{row_num})&" to $"&$J{row_num}'

                # Add formula to the specified column
                ws[f'{formula_column}{row_num}'] = formula

                # Add percentage ratio formula to column O: =(J{row}/I{row})*100
                # This calculates (new_price / old_price) * 100
                percentage_formula = f'=(J{row_num}/I{row_num})*100'
                ws[f'O{row_num}'] = percentage_formula

                formula_count += 1

            # Save the workbook
            wb.save(output_path)

            result = {
                "success": True,
                "input_file": str(input_path),
                "output_file": str(output_path),
                "output_filename": output_filename,
                "formulas_added": formula_count,
                "data_start_row": data_start_row,
                "formula_column": formula_column,
                "message": f"Successfully created formula Excel file with {formula_count} formulas"
            }

            return json.dumps(result, indent=2)

        except Exception as e:
            import json
            return json.dumps({
                "success": False,
                "error": f"Error generating formula Excel: {str(e)}"
            }, indent=2)

    def tool(self):
        """Return self for CrewAI compatibility"""
        return self


class DateExtractorInput(BaseModel):
    """Input schema for DateExtractorTool"""
    file_path: str = Field(..., description="Path to the Excel file (date will be extracted from filename)")


class DateExtractorTool(BaseTool):
    name: str = "Date Extractor from Filename"
    description: str = (
        "Extracts the effective date from TBS Price Change Summary Report filename. "
        "Expected filename format: 'TBS Price Change Summary Report - October 13th'25.xlsx' "
        "Returns the parsed date in both ISO format (YYYY-MM-DD) and display format (Month DD, YYYY)."
    )
    args_schema: type[BaseModel] = DateExtractorInput

    def _run(self, file_path: str) -> str:
        """
        Extract effective date from filename

        Args:
            file_path: Path to the Excel file (date extracted from filename)

        Returns:
            JSON string with date information
        """
        try:
            import json
            from datetime import datetime

            # Get the filename
            file_path_obj = Path(file_path)
            filename = file_path_obj.name

            # Pattern to match: Month DDth'YY or Month DD'YY
            # Examples: "October 13th'25", "November 7th'25", "January 1st'26"
            pattern = r'([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?[\'\']?(\d{2})'

            match = re.search(pattern, filename)

            if match:
                month_str = match.group(1)  # e.g., "October"
                day_str = match.group(2)     # e.g., "13"
                year_str = match.group(3)    # e.g., "25"

                # Parse the date
                date_str = f"{month_str} {day_str}, 20{year_str}"
                parsed_date = datetime.strptime(date_str, "%B %d, %Y")

                result = {
                    "success": True,
                    "source": "filename",
                    "filename": filename,
                    "effective_date_iso": parsed_date.strftime("%Y-%m-%d"),
                    "effective_date_display": parsed_date.strftime("%B %d, %Y"),
                    "raw_date_string": f"{month_str} {day_str}'{year_str}",
                    "message": f"Successfully extracted date: {parsed_date.strftime('%B %d, %Y')}"
                }

                return json.dumps(result, indent=2)
            else:
                # Fallback: use current date if pattern doesn't match
                current_date = datetime.now()
                result = {
                    "success": False,
                    "source": "fallback",
                    "filename": filename,
                    "effective_date_iso": current_date.strftime("%Y-%m-%d"),
                    "effective_date_display": current_date.strftime("%B %d, %Y"),
                    "error": "Could not extract date from filename, using current date as fallback",
                    "message": f"Using fallback date: {current_date.strftime('%B %d, %Y')}"
                }

                return json.dumps(result, indent=2)

        except Exception as e:
            import json
            from datetime import datetime

            # Return error with current date fallback
            current_date = datetime.now()
            return json.dumps({
                "success": False,
                "source": "error",
                "effective_date_iso": current_date.strftime("%Y-%m-%d"),
                "effective_date_display": current_date.strftime("%B %d, %Y"),
                "error": f"Error extracting date: {str(e)}",
                "message": f"Using fallback date: {current_date.strftime('%B %d, %Y')}"
            }, indent=2)

    def tool(self):
        """Return self for CrewAI compatibility"""
        return self
